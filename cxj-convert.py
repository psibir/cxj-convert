
# cxj-convert - a CLI CSV/XLS/XLSX/JSON converter

import argparse
import os
import csv
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class FileConverter:
    def __init__(self, input_file, output_file, output_format, delimiter=','):
        self.input_file = input_file
        self.output_file = output_file
        self.output_format = output_format
        self.delimiter = delimiter

    def convert(self):
        try:
            if self.output_format == 'csv':
                self.convert_to_csv()
            elif self.output_format == 'xls' or self.output_format == 'xlsx':
                self.convert_to_excel()
            elif self.output_format == 'json':
                self.convert_to_json()
            else:
                raise ValueError("Invalid output format")

        except Exception as e:
            print(f"Error converting {self.input_file} to {self.output_format} format: {str(e)}")

    def convert_to_csv(self):
        with open(self.input_file, 'r') as csvfile:
            reader = csv.reader(csvfile, delimiter=self.delimiter)
            rows = [row for row in reader]

        with open(self.output_file, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile, delimiter=self.delimiter)
            for row in rows:
                writer.writerow(row)

    def convert_to_excel(self):
        workbook = Workbook()

        with open(self.input_file, 'r') as csvfile:
            reader = csv.reader(csvfile, delimiter=self.delimiter)
            for row_index, row in enumerate(reader):
                for col_index, cell_value in enumerate(row):
                    col_letter = get_column_letter(col_index + 1)

        workbook.save(self.output_file)

    def convert_to_json(self):
        with open(self.input_file, 'r') as infile:
            data = csv.DictReader(infile, delimiter=self.delimiter)
            with open(self.output_file, 'w') as outfile:
                json.dump([row for row in data], outfile)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Convert files between CSV, XLS, XLSX, and JSON formats.')
    parser.add_argument('-i', '--input', type=str, help='Input file path.')
    parser.add_argument('-o', '--output', type=str, help='Output file path.')
    parser.add_argument('-f', '--format', type=str, choices=['csv', 'xls', 'xlsx', 'json'], help='Output file format.')
    parser.add_argument('-d', '--delimiter', type=str, default=',', help='Delimiter used in the CSV file.')

    args = parser.parse_args()

    if args.input and args.output and args.format:
        input_file = args.input
        output_file = args.output
        output_format = args.format
        delimiter = args.delimiter

        converter = FileConverter(input_file, output_file, output_format, delimiter)
        converter.convert()

    else:
        parser.print_help()
